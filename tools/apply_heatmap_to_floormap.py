from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ----------------------------
# ユーティリティ
# ----------------------------
def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        if isinstance(v, bool):
            return int(v)
        if isinstance(v, (int, float)):
            if math.isnan(v):
                return None
            return int(v)
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            return None
        s = s.replace(",", "")
        return int(float(s))
    except Exception:
        return None


def _read_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return json.loads(path.read_text(encoding="utf-8-sig"))


def _hex_to_rgb(h: str) -> Tuple[int, int, int]:
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    r = max(0, min(255, int(r)))
    g = max(0, min(255, int(g)))
    b = max(0, min(255, int(b)))
    return f"{r:02X}{g:02X}{b:02X}"


def _lerp(a: float, b: float, t: float) -> float:
    return a + (b - a) * t


def _lerp_color(c1: str, c2: str, t: float) -> str:
    r1, g1, b1 = _hex_to_rgb(c1)
    r2, g2, b2 = _hex_to_rgb(c2)
    r = _lerp(r1, r2, t)
    g = _lerp(g1, g2, t)
    b = _lerp(b1, b2, t)
    return _rgb_to_hex(r, g, b)


def _luminance_hex(rgb_hex_no_hash: str) -> float:
    r, g, b = _hex_to_rgb("#" + rgb_hex_no_hash)
    # 簡易輝度
    return 0.2126 * r + 0.7152 * g + 0.0722 * b


def _metric_label(metric: str) -> str:
    if metric == "diff_medals":
        return "差枚"
    if metric == "bb_rb_sum":
        return "BB+RB"
    return "最大持玉"


@dataclass
class DailyRow:
    unit: int
    max_payout: Optional[int]
    diff_medals: Optional[int]
    bb: Optional[int]
    rb: Optional[int]


def _load_daily_rows(daily_json_path: Path) -> List[DailyRow]:
    obj = _read_json(daily_json_path)

    # 形式が2種類あり得る：list か {"date":..., "rows":[...]}
    if isinstance(obj, dict) and isinstance(obj.get("rows"), list):
        rows = obj["rows"]
    elif isinstance(obj, list):
        rows = obj
    else:
        raise ValueError(f"Unsupported daily json format: {daily_json_path}")

    out: List[DailyRow] = []
    for r in rows:
        if not isinstance(r, dict):
            continue

        unit_raw = r.get("unit_no") if r.get("unit_no") is not None else r.get("machine_id")
        unit_int = _to_int(unit_raw)
        if unit_int is None:
            continue

        max_payout = _to_int(r.get("max_payout") if r.get("max_payout") is not None else r.get("max_medals"))
        diff_medals = _to_int(r.get("diff_medals") if r.get("diff_medals") is not None else r.get("diff_payout"))
        bb = _to_int(r.get("bb"))
        rb = _to_int(r.get("rb"))

        out.append(DailyRow(unit=unit_int, max_payout=max_payout, diff_medals=diff_medals, bb=bb, rb=rb))

    return out


def _pick_daily_json(repo_root: Path, date_str: Optional[str]) -> Tuple[str, Path]:
    """
    優先順位：
    1) 引数 date_str があれば docs/data/daily/<date>.json → 無ければ data/daily/<date>.json
    2) 無ければ docs/data/index.json の latest_date
    3) 無ければ docs/data/daily の最新日付ファイル
    """
    if date_str:
        p1 = repo_root / "docs" / "data" / "daily" / f"{date_str}.json"
        p2 = repo_root / "data" / "daily" / f"{date_str}.json"
        if p1.exists():
            return date_str, p1
        if p2.exists():
            return date_str, p2
        raise FileNotFoundError(f"daily json not found for date={date_str}: {p1} or {p2}")

    idx = repo_root / "docs" / "data" / "index.json"
    if idx.exists():
        obj = _read_json(idx)
        latest = obj.get("latest_date")
        if isinstance(latest, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", latest):
            p = repo_root / "docs" / "data" / "daily" / f"{latest}.json"
            if p.exists():
                return latest, p

    daily_dir = repo_root / "docs" / "data" / "daily"
    if daily_dir.exists():
        candidates = sorted([p for p in daily_dir.glob("*.json") if re.fullmatch(r"\d{4}-\d{2}-\d{2}\.json", p.name)])
        if candidates:
            latest_p = candidates[-1]
            return latest_p.stem, latest_p

    raise FileNotFoundError("Could not determine daily json. Provide --date or ensure docs/data/index.json exists.")


def _find_unit_cells(ws) -> List[Tuple[str, int]]:
    """
    フロアマップ内の「台番号セル」を探す
    3〜4桁の数字が入っているセルを台番号として扱う
    """
    hits: List[Tuple[str, int]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value).strip()
            if re.fullmatch(r"\d{3,4}", s):
                hits.append((cell.coordinate, int(s)))
    return hits


def _value_for_metric(dr: DailyRow, metric: str) -> Optional[int]:
    if metric == "diff_medals":
        return dr.diff_medals
    if metric == "bb_rb_sum":
        if dr.bb is None and dr.rb is None:
            return None
        return (dr.bb or 0) + (dr.rb or 0)
    return dr.max_payout


def _color_for_value(metric: str, v: Optional[int], vmin: float, vmax: float, max_abs: float) -> Optional[str]:
    """
    return: RGB hex (no #), e.g. 'FF0000'
    """
    if v is None:
        return None

    # 差枚は「青(マイナス)〜白(0)〜赤(プラス)」の発散スケール
    if metric == "diff_medals":
        if max_abs <= 0:
            return None
        t = max(-1.0, min(1.0, float(v) / max_abs))  # -1..1
        if t < 0:
            # blue -> white
            return _lerp_color("#2B8CBE", "#FFFFFF", t + 1.0)  # (-1->0) => (0->1)
        else:
            # white -> red
            return _lerp_color("#FFFFFF", "#DE2D26", t)

    # それ以外は「白〜赤」の連続スケール
    if vmax <= vmin:
        return _lerp_color("#FFFFFF", "#DE2D26", 0.5)

    t = (float(v) - vmin) / (vmax - vmin)
    t = max(0.0, min(1.0, t))
    return _lerp_color("#FFFFFF", "#DE2D26", t)


def _write_legend(ws, title: str, metric: str, vmin: float, vmax: float, max_abs: float) -> None:
    # 既存マップと重ならない想定で左上(A1～)に作る
    ws["A1"].value = title
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # 5段階の見本
    ws["A2"].value = "Low"
    ws["G2"].value = "High"
    ws["A2"].font = Font(bold=True)
    ws["G2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G2"].alignment = Alignment(horizontal="center", vertical="center")

    steps = 5
    for i in range(steps):
        t = i / (steps - 1)
        if metric == "diff_medals":
            # -max_abs -> +max_abs の発散
            val = int(round((-1 + 2 * t) * max_abs))
            col = _color_for_value(metric, val, vmin, vmax, max_abs)
            label = str(val)
        else:
            val = int(round(vmin + (vmax - vmin) * t))
            col = _color_for_value(metric, val, vmin, vmax, max_abs)
            label = str(val)

        cell = ws.cell(row=2, column=2 + i)  # B2..F2
        cell.value = label
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col:
            cell.fill = PatternFill("solid", fgColor=col)
            lum = _luminance_hex(col)
            cell.font = Font(color=("FFFFFF" if lum < 140 else "000000"), bold=True)

    # 少し見やすく
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--floor", default="レイトギャップ平和島　フロアマップ.xlsx", help="フロアマップxlsxのパス")
    ap.add_argument("--date", default=None, help="対象日付 YYYY-MM-DD（省略で最新）")
    ap.add_argument("--metric", default="max_payout", choices=["max_payout", "diff_medals", "bb_rb_sum"], help="色付け指標")
    ap.add_argument("--sheet", default=None, help="対象シート名（省略で先頭）")
    ap.add_argument("--out", default=None, help="出力xlsx（省略で自動命名）")
    args = ap.parse_args()

    repo_root = Path(".").resolve()

    # 日別データ決定
    date_str, daily_json_path = _pick_daily_json(repo_root, args.date)
    daily_rows = _load_daily_rows(daily_json_path)

    # unit -> value
    val_map: Dict[int, Optional[int]] = {}
    values: List[int] = []
    abs_values: List[int] = []

    for dr in daily_rows:
        v = _value_for_metric(dr, args.metric)
        val_map[dr.unit] = v
        if isinstance(v, int):
            values.append(v)
            abs_values.append(abs(v))

    if not values:
        raise RuntimeError(f"No numeric values found for metric={args.metric} in {daily_json_path}")

    vmin = float(min(values))
    vmax = float(max(values))
    max_abs = float(max(abs_values)) if abs_values else 0.0

    # フロアマップ読み込み
    floor_path = Path(args.floor)
    if not floor_path.exists():
        raise FileNotFoundError(f"Floor map not found: {floor_path}")

    wb = load_workbook(floor_path)
    ws = wb[args.sheet] if (args.sheet and args.sheet in wb.sheetnames) else wb.worksheets[0]

    # 出力は別名にして元を残す
    if args.out:
        out_path = Path(args.out)
    else:
        out_path = floor_path.with_name(f"{floor_path.stem}_heatmap_{date_str}_{args.metric}.xlsx")

    # 台番号セル探索
    unit_cells = _find_unit_cells(ws)

    # 色付け
    painted = 0
    missing = 0
    for coord, unit_int in unit_cells:
        cell = ws[coord]
        v = val_map.get(unit_int)

        col = _color_for_value(args.metric, v, vmin, vmax, max_abs)
        if col is None:
            missing += 1
            continue

        cell.fill = PatternFill("solid", fgColor=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        lum = _luminance_hex(col)
        cell.font = Font(color=("FFFFFF" if lum < 140 else "000000"), bold=True)

        painted += 1

    # 凡例
    title = f"Floor Heatmap  date={date_str}  metric={_metric_label(args.metric)}  (source: {daily_json_path.as_posix()})"
    _write_legend(ws, title, args.metric, vmin, vmax, max_abs)

    wb.save(out_path)
    print(f"[OK] saved: {out_path}")
    print(f"[INFO] painted={painted}  missing_or_null={missing}  daily={daily_json_path}")


if __name__ == "__main__":
    main()