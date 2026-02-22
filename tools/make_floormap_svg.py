# tools/make_floormap_svg.py
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_unit_cells(ws) -> List[Tuple[int, int, str]]:
    """3〜4桁の数字セルを台番号セルとして拾う (row, col, unit_str)"""
    out: List[Tuple[int, int, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = str(cell.value).strip()
            if re.fullmatch(r"\d{3,4}", s):
                out.append((cell.row, cell.column, s))
    return out


def build_merged_map(ws) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    """
    セルが結合されている場合に、(row,col)->(min_row,max_row,min_col,max_col) を引けるようにする
    """
    m: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for r in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = r.min_row, r.min_col, r.max_row, r.max_col
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                m[(rr, cc)] = (min_row, max_row, min_col, max_col)
    return m


def esc(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="レイトギャップ平和島　フロアマップ.xlsx", help="入力xlsx")
    ap.add_argument("--sheet", default="Sheet1", help="シート名")
    ap.add_argument("--out", default="docs/assets/floormap.svg", help="出力svg")
    ap.add_argument("--cell-w", type=float, default=34.0, help="1セルの幅(px)")
    ap.add_argument("--cell-h", type=float, default=22.0, help="1セルの高さ(px)")
    ap.add_argument("--font-size", type=float, default=10.0, help="文字サイズ(px)")
    ap.add_argument("--stroke", default="#1c2433", help="枠線色")
    ap.add_argument("--stroke-width", type=float, default=1.2, help="枠線太さ")
    args = ap.parse_args()

    xlsx = Path(args.inp)
    if not xlsx.exists():
        raise FileNotFoundError(f"not found: {xlsx}")

    wb = load_workbook(xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {args.sheet} (exists: {wb.sheetnames})")
    ws = wb[args.sheet]

    unit_cells = find_unit_cells(ws)
    if not unit_cells:
        raise RuntimeError("台番号セル(3〜4桁)が見つかりません")

    merged_map = build_merged_map(ws)

    min_row = min(r for r, c, u in unit_cells)
    max_row = max(r for r, c, u in unit_cells)
    min_col = min(c for r, c, u in unit_cells)
    max_col = max(c for r, c, u in unit_cells)

    width = (max_col - min_col + 1) * args.cell_w
    height = (max_row - min_row + 1) * args.cell_h

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # --- SVG生成 ---
    lines: List[str] = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'viewBox="0 0 {width:.0f} {height:.0f}" '
        f'preserveAspectRatio="xMidYMid meet">'
    )

    # 内部CSS（最低限）
    lines.append("<style>")
    lines.append(".machine-rect{fill:#FFFFFF;}")
    lines.append(".machine-text{fill:#000000;font-family:system-ui,-apple-system,Segoe UI,Roboto,'Noto Sans JP',sans-serif;}")
    lines.append(".machine{cursor:pointer;}")
    lines.append("</style>")

    # 台番号セルだけを描画（空白セルは描画しない）
    for (r, c, unit_str) in unit_cells:
        # 結合セル対応
        if (r, c) in merged_map:
            mr1, mr2, mc1, mc2 = merged_map[(r, c)]
        else:
            mr1, mr2, mc1, mc2 = r, r, c, c

        x = (mc1 - min_col) * args.cell_w
        y = (mr1 - min_row) * args.cell_h
        w = (mc2 - mc1 + 1) * args.cell_w
        h = (mr2 - mr1 + 1) * args.cell_h

        # unit は "0729" ではなく "729" でID化（先頭0を落として一致しやすく）
        unit_id = str(int(unit_str))  # "0729" -> "729"
        gid = f"unit-{unit_id}"

        # 中央配置
        cx = x + w / 2.0
        cy = y + h / 2.0

        lines.append(f'<g id="{gid}" class="machine" data-unit="{unit_id}">')
        lines.append(
            f'<rect class="machine-rect" x="{x:.1f}" y="{y:.1f}" width="{w:.1f}" height="{h:.1f}" '
            f'stroke="{esc(args.stroke)}" stroke-width="{args.stroke_width:.1f}" />'
        )
        lines.append(
            f'<text class="machine-text" x="{cx:.1f}" y="{cy:.1f}" font-size="{args.font_size:.1f}" '
            f'text-anchor="middle" dominant-baseline="central">{esc(unit_id)}</text>'
        )
        lines.append("</g>")

    lines.append("</svg>")

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[OK] wrote: {out_path}")
    print(f"[INFO] units={len(unit_cells)}  range={get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}")


if __name__ == "__main__":
    main()